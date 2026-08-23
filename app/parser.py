from __future__ import annotations

import json
import re
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.models import Transaction, money

EXCEL_EPOCH = datetime(1899, 12, 30)
DEFAULTS_PATH = Path(__file__).resolve().parent.parent / "config" / "defaults.json"


def load_defaults() -> dict[str, Any]:
    return json.loads(DEFAULTS_PATH.read_text())


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = int(value)
        if serial > 20000:
            return (EXCEL_EPOCH + timedelta(days=serial)).date()
    text = _norm(value)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return money(value)
    text = _norm(value).replace("$", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return money(float(text))
    except ValueError:
        return 0.0


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    return _norm(value)


def _cell_map(ws) -> dict[str, Any]:
    labels: dict[str, list[tuple[int, int, Any]]] = {}
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=6):
        for cell in row:
            key = _norm(cell.value).upper()
            if key:
                labels.setdefault(key, []).append((cell.row, cell.column, cell.value))
    return labels


def _value_right(ws, row: int, col: int) -> Any:
    for offset in (1, 2, 3, 4, 5):
        cell = ws.cell(row=row, column=col + offset)
        if cell.value not in (None, ""):
            return cell.value
    return None


def _first_label(labels: dict[str, list[tuple[int, int, Any]]], *names: str) -> tuple[int, int] | None:
    for name in names:
        hits = labels.get(name.upper())
        if hits:
            return hits[0][0], hits[0][1]
    return None


def _sale_checked(xlsx_path: Path) -> bool:
    """Read SALE vs LEASE form-control checkboxes from the xlsx drawing."""
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = [n for n in zf.namelist() if n.endswith(".vml") or "vmlDrawing" in n]
            if not names:
                return True
            xml = zf.read(names[0]).decode("utf-8", errors="ignore")
        # First two ClientData checkboxes are SALE and LEASE.
        blocks = re.findall(
            r"<x:ClientData ObjectType=\"Checkbox\">(.*?)</x:ClientData>",
            xml,
            flags=re.S,
        )
        if not blocks:
            return True
        sale_checked = "Checked" in blocks[0] or "<x:Checked>1</x:Checked>" in blocks[0]
        return sale_checked
    except Exception:
        return True


def apply_overrides(tx: Transaction, overrides: dict[str, Any] | None = None) -> Transaction:
    overrides = overrides or {}
    for key, value in overrides.items():
        if not hasattr(tx, key):
            continue
        if key == "gross_commission":
            parsed_gross = _to_number(value)
            tx._gross_commission_override = parsed_gross
            tx.total_commission = parsed_gross
            continue
        current = getattr(tx, key)
        if isinstance(current, float):
            setattr(tx, key, _to_number(value))
        elif isinstance(current, date) or key in {"today", "close_date"}:
            setattr(tx, key, _to_date(value))
        elif isinstance(current, bool):
            setattr(tx, key, str(value).lower() in {"1", "true", "yes", "sale"})
        else:
            setattr(tx, key, _to_text(value))
    return tx


def transaction_from_overrides(overrides: dict[str, Any] | None = None) -> Transaction:
    defaults = load_defaults()
    brokerage = defaults.get("brokerage", {})
    broker = defaults.get("broker", {})
    tx = Transaction(
        agent_payee_address="",
        broker_name=broker.get("name", ""),
        brokerage_name=brokerage.get("name", ""),
        brokerage_mail_address=", ".join(
            p for p in [brokerage.get("address_line1", ""), brokerage.get("address_line2", "")] if p
        ),
        brokerage_email=brokerage.get("email", ""),
        brokerage_phone=brokerage.get("phone", ""),
    )
    return apply_overrides(tx, overrides)


def parse_rop_xlsx(path: Path, overrides: dict[str, Any] | None = None) -> Transaction:
    defaults = load_defaults()
    overrides = overrides or {}
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    labels = _cell_map(ws)

    def labeled(*names: str) -> Any:
        loc = _first_label(labels, *names)
        if not loc:
            return None
        return _value_right(ws, loc[0], loc[1])

    def labeled_nth(name: str, index: int) -> Any:
        hits = labels.get(name.upper(), [])
        if len(hits) <= index:
            return None
        row, col, _ = hits[index]
        return _value_right(ws, row, col)

    selling_agent = _to_text(labeled("SELLING AGENT"))
    brokerage = defaults.get("brokerage", {})
    broker = defaults.get("broker", {})
    payees = defaults.get("agent_payee_addresses", {})

    tx = Transaction(
        today=_to_date(labeled("TODAY'S DATE", "TODAYS DATE")),
        close_date=_to_date(labeled("CLOSING DATE")),
        is_sale=_sale_checked(path),
        property_address=_to_text(labeled("PROPERTY ADDRESS")),
        sale_price=_to_number(labeled("SALE PRICE")),
        mls=_to_text(labeled("MLS#", "MLS")),
        seller=_to_text(labeled("SELLER/LANDLORD", "SELLER")),
        buyer=_to_text(labeled("BUYER/TENANT", "BUYER")),
        listing_broker=_to_text(labeled("LISTING BROKER")),
        listing_broker_phone=_to_text(labeled_nth("PHONE #", 0) or labeled_nth("PHONE#", 0)),
        listing_agent=_to_text(labeled("LISTING AGENT")),
        listing_agent_phone=_to_text(labeled_nth("PHONE #", 1) or labeled_nth("PHONE#", 1)),
        listing_agent_email=_to_text(labeled_nth("EMAIL", 0)),
        selling_broker=_to_text(labeled("SELLING BROKER")),
        selling_broker_phone=_to_text(labeled_nth("PHONE #", 2) or labeled_nth("PHONE#", 2)),
        selling_broker_address=_to_text(labeled_nth("ADDRESS", 1) or labeled("ADDRESS")),
        selling_agent=selling_agent,
        selling_agent_phone=_to_text(labeled_nth("PHONE #", 3) or labeled_nth("PHONE#", 3)),
        selling_agent_email=_to_text(labeled_nth("EMAIL", 1)),
        title_company=_to_text(labeled("TITLE CO./ADDRESS", "TITLE CO.", "TITLE COMPANY")),
        closer=_to_text(labeled("CLOSER")),
        closer_email=_to_text(labeled_nth("EMAIL", 2)),
        closer_phone=_to_text(labeled_nth("PHONE #", 4) or labeled_nth("PHONE#", 4)),
        escrow_no=_to_text(labeled("GF #", "GF#", "ESCROW NO", "ESCROW #")),
        broker_process_fees=_to_number(ws["F29"].value),
        broker_other_fees=_to_number(ws["F30"].value),
        selling_agent_commission=_to_number(ws["F31"].value),
        listing_agent_commission=_to_number(ws["F32"].value),
        agent_contribution=_to_number(ws["F33"].value),
        other_fees=_to_number(ws["F34"].value),
        total_commission=_to_number(ws["F35"].value),
        agent_payee_address=payees.get(selling_agent, ""),
        broker_name=broker.get("name", ""),
        brokerage_name=brokerage.get("name", ""),
        brokerage_mail_address=", ".join(
            p for p in [brokerage.get("address_line1", ""), brokerage.get("address_line2", "")] if p
        ),
        brokerage_email=brokerage.get("email", ""),
        brokerage_phone=brokerage.get("phone", ""),
    )

    if not tx.agent_payee_address:
        tx.warnings.append(
            f"No payee mailing address on file for selling agent “{selling_agent or 'unknown'}”. "
            "Add it in the form or in config/defaults.json."
        )

    return apply_overrides(tx, overrides)
